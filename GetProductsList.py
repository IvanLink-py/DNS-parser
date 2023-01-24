import sys

from bs4 import BeautifulSoup
from selenium import webdriver
import openpyxl

start_pos = 4
browser = webdriver.Chrome()

def product_attr_2_ref(attr: str):
    return f"https://www.dns-shop.ru/product/{attr[:8]}{attr[9:13]}{attr[31:35]}"


catalog_pages = {
    "CPU": [
        f"https://www.dns-shop.ru/"
        f"catalog/17a899cd16404e77/processory/?stock=now-today-tomorrow-later-out_of_stock&p={i}"
        for i in range(1, 33)],
    "MotherBoard": [
        f"https://www.dns-shop.ru/"
        f"catalog/17a89a0416404e77/materinskie-platy/?stock=now-today-tomorrow-later-out_of_stock&p={i}"
        for i in range(1, 62)],
    "GPU": [
        f"https://www.dns-shop.ru/"
        f"catalog/17a89aab16404e77/videokarty/?stock=now-today-tomorrow-later-out_of_stock&p={i}"
        for i in range(1, 66)],
    "RAM": [
        f"https://www.dns-shop.ru/"
        f"catalog/17a89a3916404e77/operativnaya-pamyat-dimm/?stock=now-today-tomorrow-later-out_of_stock&p={i}"
        for i in range(1, 111)],
    "PowerSupply": [
        f"https://www.dns-shop.ru/"
        f"catalog/17a89c2216404e77/bloki-pitaniya/?stock=now-today-tomorrow-later-out_of_stock&p={i}"
        for i in range(1, 78)],
    "HDD": [
        f"https://www.dns-shop.ru/"
        f"catalog/17a8914916404e77/zhestkie-diski-35/?stock=now-today-tomorrow-later-out_of_stock&p={i}"
        for i in range(1, 19)],
    "SSD": [
        f"https://www.dns-shop.ru/"
        f"catalog/8a9ddfba20724e77/ssd-nakopiteli/?stock=now-today-tomorrow-later-out_of_stock&p={i}"
        for i in range(1, 26)]
}


def get_page_products(soup: BeautifulSoup):
    return [(i['data-product'], i) for i in soup.find_all(class_="catalog-product")]


def get_products_from_catalog_ref(href: str):
    browser.get(href)
    html = browser.page_source
    soup = BeautifulSoup(html, 'lxml')

    products_data = get_page_products(soup)
    return map(lambda t:
               (product_attr_2_ref(t[0]),
                t[1].find_next(class_='catalog-product__name').text),
               products_data)


def write_urls(workbook, sheet_name, data):
    wb = openpyxl.load_workbook(workbook)
    sheet = wb[sheet_name]

    free_row = start_pos
    while sheet[f"B{free_row}"].value is not None:
        free_row += 1
        if free_row > 5000:
            raise Exception()

    for i, data in enumerate(data):
        sheet.cell(row=free_row + i, column=2, value=data[0])
        sheet.cell(row=free_row + i, column=3, value=data[1])

    wb.save(filename=workbook)


def main(limit=-1, start=0):
    for cat in list(catalog_pages.keys()):
        for cat_page_ref in catalog_pages[cat]:

            if start == 0:

                data = get_products_from_catalog_ref(cat_page_ref)
                write_urls("Export.xlsx", "Export", data)
                print(f"{limit} - {cat_page_ref}")

            else:
                start -= 1

            limit -= 1
            if limit == 0:
                sys.exit(0)


if __name__ == '__main__':
    main()
