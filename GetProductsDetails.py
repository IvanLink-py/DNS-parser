from bs4 import BeautifulSoup
import openpyxl
from GetProductsList import start_pos, catalog_pages
from GetProductsList import browser
from pprint import pp


def get_ref_by_cat(workbook, cat):
    start = start_pos

    refs_by_cats = []

    wb = openpyxl.load_workbook(workbook, read_only=True)

    index = 0

    sheet = wb[cat]

    while True:
        value = sheet.cell(column=2, row=start + index).value

        if value is None:
            break

        refs_by_cats.append((start + index, value))

        index += 1

    return refs_by_cats


def get_details_by_ref(ref):
    browser.get(ref)
    html = browser.page_source
    soup = BeautifulSoup(html, 'lxml')

    specs = {}

    try:
        for spec in soup.find_all(class_="product-characteristics__spec"):
            spec_name = spec.find_next(class_="product-characteristics__spec-title").text.strip()
            spec_val = spec.find_next(class_="product-characteristics__spec-value").text.strip()

            specs[spec_name] = spec_val
    except AttributeError:
        return {}

    return specs


def write_details(workbook, sheet_name, row, details):
    def get_spec_column(shet, spec_):
        i = 4
        while True:
            value = shet.cell(row=3, column=i).value
            if value is None or value == spec_:
                return i
            i += 1

            if i > 5000:
                raise GeneratorExit()

    wb = openpyxl.load_workbook(workbook)
    sheet = wb[sheet_name]

    for spec in details:
        column = get_spec_column(sheet, spec)

        sheet.cell(row=3, column=column, value=spec)
        sheet.cell(row=row, column=column, value=details[spec])

    wb.save(workbook)


def main():
    skip = 34
    for cat in catalog_pages:
        refs = get_ref_by_cat("Export.xlsx", cat)
        for row, ref in refs:

            if skip > 0:
                print(cat, row, f'Skip {skip}')
                skip -= 1
                continue

            specs = get_details_by_ref(ref)
            repeat = 0
            while specs == {} and repeat < 5:
                specs = get_details_by_ref(ref)
                repeat += 1

            if specs == {}:
                print(cat, row, 'Error')
                continue

            write_details("Export.xlsx", cat, row, specs)

            print(cat, row, ref, specs)


if __name__ == '__main__':
    main()
